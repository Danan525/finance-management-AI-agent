"""财务报表 Excel 导出（openpyxl）：封面 + 三张报表 + 勾稽校验 + 取数/凭证审计轨迹。

对齐报表计划：**勾稽不通过不出表**（导出前由 service 校验 can_issue，本模块只负责排版）。
每个报表数字都能向下追溯：'科目余额'页给"科目→报表行"取数轨迹，'分录明细'页给凭证轨迹。
金额显示用 float + 数字格式（与 invoice-excel writer 一致）；底层计算全程 Decimal。
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ledger import accounts as A
from ledger.engine import ZERO
from ledger.service import load_ledger
from . import service as rsvc

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
BOLD = Font(bold=True)
OK_FILL = PatternFill("solid", fgColor="E2EFDA")       # 绿：勾稽通过
BAD_FILL = PatternFill("solid", fgColor="F8CBAD")      # 橙红：勾稽未过
SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")
MONEY_FMT = "#,##0.00"
RIGHT = Alignment(horizontal="right")


def _f(v):
    return float(v) if isinstance(v, Decimal) else float(Decimal(str(v)))


def _money(ws, row, col, value):
    c = ws.cell(row=row, column=col)
    c.value = _f(value)
    c.number_format = MONEY_FMT
    c.alignment = RIGHT
    return c


def _hrow(ws, row, headers, start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT


def build_reports_workbook(out_path: Path, generated_at: str = "") -> Path:
    rpt = rsvc.generate()
    wb = Workbook()

    _cover(wb.active, rpt, generated_at)
    _income(wb.create_sheet("利润表 Income Statement"), rpt["income_statement"])
    _balance(wb.create_sheet("资产负债表 Balance Sheet"), rpt["balance_sheet"])
    _cashflow(wb.create_sheet("现金流量表 Cash Flow"), rpt["cash_flow_statement"])
    _checks(wb.create_sheet("勾稽校验 Reconciliation"), rpt["checks"])
    _account_balances(wb.create_sheet("科目余额 取数轨迹"))
    _entries(wb.create_sheet("分录明细 凭证轨迹"))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return out_path


def _cover(ws, rpt, generated_at):
    ws.title = "封面 Cover"
    ws.cell(1, 1, "财务报表 Financial Statements").font = TITLE_FONT
    b = rpt["basis"]
    rows = [
        ("编制基础 Framework", b["framework"]),
        ("功能货币 Currency", b["currency"]),
        ("报告期间 Period", b["period"]),
        ("生成时间 Generated", generated_at or "-"),
        ("勾稽结论 Reconciliation", "通过 · 可出表" if rpt["can_issue"] else "未通过 · 草稿"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(i, 1, k).font = BOLD
        ws.cell(i, 2, v)
    ws.cell(9, 1, "本报表由总账已过账分录自动生成；各数字可在'科目余额'与'分录明细'页向下追溯。")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40


def _income(ws, istmt):
    _hrow(ws, 1, ["项目 Item", "金额 Amount"])
    for i, l in enumerate(istmt["lines"], start=2):
        c = ws.cell(i, 1, l["label"])
        mc = _money(ws, i, 2, l["amount"])
        if l["kind"] in ("subtotal", "total"):
            c.font = BOLD
            mc.font = BOLD
        if l["kind"] == "subtotal":
            c.fill = SECTION_FILL
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 18


def _section(ws, row, title):
    c = ws.cell(row, 1, title)
    c.font = BOLD
    c.fill = SECTION_FILL
    ws.cell(row, 2).fill = SECTION_FILL


def _balance(ws, bs):
    _hrow(ws, 1, ["项目 Item", "金额 Amount"])
    r = 2
    _section(ws, r, "资产 Assets"); r += 1
    for x in bs["assets"]:
        ws.cell(r, 1, "  " + x["label"]); _money(ws, r, 2, x["amount"]); r += 1
    ws.cell(r, 1, "资产合计 Total Assets").font = BOLD; _money(ws, r, 2, bs["assets_total"]).font = BOLD; r += 1
    _section(ws, r, "负债 Liabilities"); r += 1
    for x in bs["liabilities"]:
        ws.cell(r, 1, "  " + x["label"]); _money(ws, r, 2, x["amount"]); r += 1
    ws.cell(r, 1, "负债合计 Total Liabilities").font = BOLD; _money(ws, r, 2, bs["liabilities_total"]).font = BOLD; r += 1
    _section(ws, r, "权益 Equity"); r += 1
    for x in bs["equity"]:
        ws.cell(r, 1, "  " + x["label"]); _money(ws, r, 2, x["amount"]); r += 1
    ws.cell(r, 1, "权益合计 Total Equity").font = BOLD; _money(ws, r, 2, bs["equity_total"]).font = BOLD; r += 1
    ws.cell(r, 1, "负债 + 权益 Total Liab. & Equity").font = BOLD
    _money(ws, r, 2, bs["liab_equity_total"]).font = BOLD; r += 1
    ws.cell(r, 1, "平衡校验 Balanced").font = BOLD
    bc = ws.cell(r, 2, "✓ 平衡" if bs["balanced"] else f"✗ 差 {bs['diff']}")
    bc.fill = OK_FILL if bs["balanced"] else BAD_FILL
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18


def _cashflow(ws, cf):
    _hrow(ws, 1, ["项目 Item", "金额 Amount"])
    r = 2
    for l in cf["lines"]:
        ws.cell(r, 1, "  " + l["label"]); _money(ws, r, 2, l["amount"]); r += 1
    ws.cell(r, 1, "现金净增加 Net Change").font = BOLD; _money(ws, r, 2, cf["net_change"]).font = BOLD; r += 1
    ws.cell(r, 1, "期初现金 Opening"); _money(ws, r, 2, cf["opening"]); r += 1
    ws.cell(r, 1, "期末现金 Ending").font = BOLD; _money(ws, r, 2, cf["ending"]).font = BOLD; r += 1
    ws.cell(r, 1, "与货币资金勾稽 E3").font = BOLD
    ec = ws.cell(r, 2, "✓ 一致" if cf["e3_ok"] else f"✗ 差 {cf['e3_diff']}")
    ec.fill = OK_FILL if cf["e3_ok"] else BAD_FILL
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18


def _checks(ws, ck):
    _hrow(ws, 1, ["勾稽项 Check", "说明", "结果", "差异/明细"])
    r = 2
    for key in ("E1_balance_sheet_balanced", "E3_cash_tie", "E6_all_classified"):
        c = ck.get(key)
        if not c:
            continue
        ws.cell(r, 1, key)
        ws.cell(r, 2, c.get("desc", ""))
        rc = ws.cell(r, 3, "✓ 通过" if c["ok"] else "✗ 未过")
        rc.fill = OK_FILL if c["ok"] else BAD_FILL
        detail = c.get("diff", "") or "、".join(c.get("unclassified", []))
        ws.cell(r, 4, detail)
        r += 1
    ws.cell(r + 1, 1, "可出表 can_issue").font = BOLD
    fc = ws.cell(r + 1, 3, "是" if ck["can_issue"] else "否")
    fc.fill = OK_FILL if ck["can_issue"] else BAD_FILL
    for col, w in (("A", 30), ("B", 40), ("C", 12), ("D", 40)):
        ws.column_dimensions[col].width = w


def _account_balances(ws):
    """取数轨迹：每个科目的借/贷/净额 + 报表行归属（未归类标红）。"""
    _hrow(ws, 1, ["科目 Account", "借方合计", "贷方合计", "净额(借-贷)", "报表行归属"])
    led = load_ledger()
    _dr, _cr, rows = led.trial_balance()
    r = 2
    for account, dr, cr in rows:
        ws.cell(r, 1, account)
        _money(ws, r, 2, dr); _money(ws, r, 3, cr); _money(ws, r, 4, dr - cr)
        token = A.report_line(account)
        tc = ws.cell(r, 5, token or "（未归类）")
        if not token:
            tc.fill = BAD_FILL
        r += 1
    for col, w in (("A", 40), ("B", 14), ("C", 14), ("D", 14), ("E", 34)):
        ws.column_dimensions[col].width = w


def _entries(ws):
    """凭证轨迹：全部已过账/红冲分录的逐行明细。"""
    _hrow(ws, 1, ["凭证号", "日期", "状态", "摘要", "科目", "借方", "贷方", "活动"])
    from ledger import store
    r = 2
    for e in store.entries_for_balance(limit=1000000):
        act = getattr(e, "activity", "") or ""
        for l in e.lines:
            ws.cell(r, 1, e.entry_no); ws.cell(r, 2, e.date); ws.cell(r, 3, e.status)
            ws.cell(r, 4, e.memo); ws.cell(r, 5, l.account)
            if l.debit and l.debit != ZERO:
                _money(ws, r, 6, l.debit)
            if l.credit and l.credit != ZERO:
                _money(ws, r, 7, l.credit)
            ws.cell(r, 8, A.ACTIVITY_LABEL.get(act, act))
            r += 1
    for col, w in (("A", 14), ("B", 12), ("C", 10), ("D", 30), ("E", 34),
                   ("F", 14), ("G", 14), ("H", 20)):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
