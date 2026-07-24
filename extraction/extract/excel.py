"""Excel(.xlsx) 发票解析 + 自渲染（保真显示 + 字段高亮 + 长文本换行）。

**不走 fitz**：fitz 抽 xlsx 把日期渲染成序列号、丢格式。改用 **openpyxl**：
① 读结构化单元格（日期=datetime→ISO；金额按 number_format 货币格式带两位小数；数量整数保留）；
② 按"列→x、行→y"合成统一坐标布局（**长文本单元格自动换行**，行高随之增加），复用通用解析；
③ 用 Pillow 按**同一坐标**自渲染 PNG，保留加粗/字色/填充色，并据同一坐标回填字段 bbox
   → 审核页字段↔原件双向高亮。长内容换行后不再溢出画面，配合前端横向滚动可完整核对。
"""
from __future__ import annotations

import datetime
import re
from collections import defaultdict
from copy import copy as _copy
from pathlib import Path
from typing import List, Optional, Tuple

from .pdf_text import PdfDoc, pdfdoc_from_word_tuples

_INV_ANCHOR = re.compile(r"invoice\s*(no|number|#)", re.IGNORECASE)

_COLW = 130.0      # 列间距 pt
_ROWH = 18.0       # 行高 pt（单行）
_LINEH = 15.0      # 换行行距 pt
_PAD = 12.0        # 页边距 pt
_CHARW = 6.0       # cell 分组用的保守字宽
_MAXTW = 520.0     # 无右邻单元格时，长文本换行的最大宽度 pt
_EXCEL_EXTS = (".xlsx", ".xlsm")
# 字体目录优先级：**随包字体（跨平台一致，首选）** → Linux 系统 DejaVu（兜底）。
# 不能只写系统路径：Mac/Windows 无此路径会导致 truetype 失败、退回 PIL 位图字体（字号失效、
# 大小全乱："该大的小、该小的大"）。随包字体保证任何平台渲染都和 Linux 一致。
_FONT_DIRS = [str(Path(__file__).resolve().parent / "fonts"),
              "/usr/share/fonts/truetype/dejavu"]
_MIN_IMG_AREA = 200_000          # 像素面积下限：≈ 500×400 以上才算"整页发票图"，滤掉 logo/banner/装饰图
_MIN_TEXT_CHARS = 100            # 单元格可提取文本少于此 → 视为"图片形式发票"
_RASTER = {"PNG": ".png", "JPEG": ".jpg", "BMP": ".bmp", "TIFF": ".tif", "GIF": ".gif"}


def is_excel(path) -> bool:
    return Path(path).suffix.lower() in _EXCEL_EXTS


def extract_images(path) -> List[Tuple[bytes, str]]:
    """提取 xlsx 内嵌的**位图发票图**（过滤过小的 logo/装饰）→ [(bytes, ext), ...]。"""
    import io
    import openpyxl
    from PIL import Image
    out: List[Tuple[bytes, str]] = []
    for ws in openpyxl.load_workbook(str(path)).worksheets:
        for im in getattr(ws, "_images", []):
            try:
                data = im._data() if callable(getattr(im, "_data", None)) else None
                pic = Image.open(io.BytesIO(data))
                w, h = pic.size
            except Exception:
                continue
            ext = _RASTER.get((pic.format or "").upper())
            if ext and w * h >= _MIN_IMG_AREA:
                out.append((data, ext))
    return out


_INV_MARKER = re.compile(
    r"invoice\s*(no|number|#)|bill\s*no|total|amount\s+due|grand\s+total|subtotal|"
    r"发票|账单|合计|总额|总计|金额|应付|税额", re.IGNORECASE)


def is_image_form(path) -> bool:
    """是否"图片形式发票"：含 ≥1 张位图发票图，且 单元格文本很少 或 文本里无发票要素
    （真发票在图里）——避免"图+大段文字"的 xlsx 漏掉图里的发票。"""
    if len(extract_images(path)) < 1:
        return False
    cells = " ".join(t for (_c, _r, t, *_s) in _read_cells(path))
    return len(cells) < _MIN_TEXT_CHARS or not _INV_MARKER.search(cells)


def _fmt(value, number_format: str) -> Optional[str]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        nf = (number_format or "").upper()
        if any(s in nf for s in ("$", "€", "£", "0.00", "#,##0", "USD", "EUR", "GBP")):
            return f"{float(value):,.2f}"
        if float(value) == int(value):
            return str(int(value))
        return ("%f" % value).rstrip("0").rstrip(".")
    return str(value).strip() or None


def _rgb(color) -> Optional[Tuple[int, int, int]]:
    try:
        v = getattr(color, "rgb", None)
        if isinstance(v, str) and len(v) >= 6 and not v.upper().startswith("00"):
            h = v[-6:]
            return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
    except Exception:
        pass
    return None


def _read_cells(path) -> list:
    """读 xlsx 活动表 → [(col,row,text,bold,size,color,fill)]。"""
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb.active
    out = []
    for row in ws.iter_rows():
        for c in row:
            t = _fmt(c.value, c.number_format)
            if not t:
                continue
            f = c.font
            out.append((c.column, c.row, t, bool(f and f.bold),
                        float(f.sz) if f and f.sz else 10.0,
                        _rgb(f.color) if f else None,
                        _rgb(c.fill.fgColor) if c.fill and c.fill.patternType == "solid" else None))
    return out


def _wrap(text: str, max_w: float, size: float) -> List[str]:
    """按估算字宽把文本折行到 max_w 内（超长单词硬切）。"""
    cw = max(size * 0.55, 4.0)
    maxc = max(6, int(max_w / cw))
    lines, cur = [], ""
    for w in text.split():
        while len(w) > maxc:                       # 超长单词硬切
            if cur:
                lines.append(cur); cur = ""
            lines.append(w[:maxc]); w = w[maxc:]
        if cur and len(cur) + 1 + len(w) > maxc:
            lines.append(cur); cur = w
        else:
            cur = (cur + " " + w) if cur else w
    if cur:
        lines.append(cur)
    return lines or [text]


def _place(path):
    """统一布局：解析与渲染共用。返回 (placed_cells, page_w, page_h)。
    placed cell: dict(col,row,text,bold,size,color,fill,x0,y0,x1,y1,lines)。
    """
    raw = _read_cells(path)
    byrow = defaultdict(list)
    for c in raw:
        byrow[c[1]].append(c)
    # 每个 cell 的可用宽度 + 换行
    wrapped, rowlines = {}, {}
    for r, cells in byrow.items():
        cells.sort(key=lambda c: c[0])
        for i, c in enumerate(cells):
            col, size = c[0], c[4]
            x0 = _PAD + (col - 1) * _COLW
            if i + 1 < len(cells):                 # 有右邻 → 折行到列内，避免重叠
                avail = max(_COLW - 8, _PAD + (cells[i + 1][0] - 1) * _COLW - x0 - 8)
            else:
                avail = _MAXTW
            wrapped[(col, r)] = (_wrap(c[2], avail, size), avail)
        rowlines[r] = max(len(wrapped[(c[0], r)][0]) for c in cells)
    # 累计 y（换行多出的行把后续行下推）
    rowtop, extra = {}, 0
    for r in sorted(byrow):
        rowtop[r] = _PAD + (r - 1) * _ROWH + extra * _LINEH
        extra += rowlines[r] - 1
    placed, maxx, maxy = [], 0.0, 0.0
    for r, cells in byrow.items():
        for c in cells:
            col, size = c[0], c[4]
            lines, avail = wrapped[(col, r)]
            x0 = _PAD + (col - 1) * _COLW
            y0 = rowtop[r]
            tw = max((len(ln) * size * 0.55 for ln in lines), default=0)
            w = min(tw, avail) if len(lines) > 1 else tw
            h = len(lines) * _LINEH if len(lines) > 1 else _ROWH - 4
            x1, y1 = x0 + w, y0 + h
            placed.append({"col": col, "row": r, "text": c[2], "bold": c[3], "size": size,
                           "color": c[5], "fill": c[6], "x0": x0, "y0": y0, "x1": x1, "y1": y1,
                           "lines": lines})
            maxx, maxy = max(maxx, x1), max(maxy, y1)
    return placed, maxx + _PAD, maxy + _PAD


def excel_to_pdfdoc(path) -> PdfDoc:
    """合成 PdfDoc：lines（解析用 cell 分组）+ words_geom（高亮用）+ page_sizes。"""
    placed, page_w, page_h = _place(path)
    parse_words, geom, rows_text = [], [], defaultdict(list)
    for c in placed:
        x0, y0, t = c["x0"], c["y0"], c["text"]
        parse_words.append((x0, y0, x0 + min(len(t) * _CHARW, _COLW - 30), y0 + _ROWH - 4, t))
        geom.append((0, x0, y0, c["x1"], c["y1"], t))
        rows_text[c["row"]].append((c["col"], t))
    full = "\n".join(" ".join(t for _, t in sorted(rows_text[r])) for r in sorted(rows_text))
    doc = pdfdoc_from_word_tuples(parse_words, full_text=full)
    doc.words_geom = geom
    doc.page_sizes = [[round(page_w, 1), round(page_h, 1)]]
    return doc


# ---- 多发票检测与物理拆分（与 PDF 同逻辑：先拆成一文件一发票，再走单张路径）----

def _rows_text(ws) -> dict:
    """{行号: 该行各非空单元格文本拼接}（用于按内容定位发票边界）。"""
    out = {}
    for row in ws.iter_rows():
        cells = [(c.column, _fmt(c.value, c.number_format)) for c in row]
        cells = [(col, t) for col, t in cells if t]
        if cells:
            out[row[0].row] = " ".join(t for _, t in sorted(cells))
    return out


def invoice_units(path) -> Optional[List[Tuple]]:
    """检测 xlsx 内多张发票 → 单元列表 [(sheet, r0, r1), ...]（len≥2 才拆）；单张/不确定返回 None。

    与 PDF 同思路（TOTAL DUE 数=发票数 + 内容锚点定边界 + 每段恰含 1 个 TOTAL 的完整性校验）：
    ① 每个工作表若恰含 1 张发票 → 整表为一单元；② 单表内多张（多个 TOTAL）→ 按 Invoice No 锚点行
    切段，每段须恰含 1 个 TOTAL，否则判不可靠、返回 None（绝不误拆）。r0/r1=None 表示整表。
    """
    from ..parse.template_rules import count_total_markers
    import openpyxl
    wb = openpyxl.load_workbook(str(path), data_only=True)
    cand: List[Tuple] = []
    for ws in wb.worksheets:
        rt = _rows_text(ws)
        if not rt:
            continue
        n = count_total_markers("\n".join(rt[r] for r in sorted(rt)))
        if n == 0:
            continue                                  # 该表无发票
        if n == 1:
            cand.append((ws.title, None, None))
            continue
        anchors = [r for r in sorted(rt) if _INV_ANCHOR.search(rt[r])]
        if len(anchors) != n:
            return None                               # 边界切不清楚 → 不拆
        bounds = anchors + [max(rt) + 1]
        for i in range(len(anchors)):
            r0, r1 = bounds[i], bounds[i + 1] - 1
            band = "\n".join(rt[r] for r in sorted(rt) if r0 <= r <= r1)
            if count_total_markers(band) != 1:
                return None
            cand.append((ws.title, r0, r1))
    return cand if len(cand) >= 2 else None


def _copy_range(src_ws, dst_ws, r0, r1) -> None:
    """把源表 [r0,r1] 行（None=整表）连样式/数字格式/列宽复制到目标表（行号重映射从 1 起）。"""
    rmin = r0 or src_ws.min_row
    rmax = r1 or src_ws.max_row
    off = rmin - 1
    for row in src_ws.iter_rows(min_row=rmin, max_row=rmax):
        for c in row:
            if c.value is None:
                continue
            d = dst_ws.cell(row=c.row - off, column=c.column, value=c.value)
            d.number_format = c.number_format
            if c.has_style:
                d.font, d.fill, d.alignment = _copy(c.font), _copy(c.fill), _copy(c.alignment)
    for col, dim in src_ws.column_dimensions.items():
        if dim.width:
            dst_ws.column_dimensions[col].width = dim.width


def split_xlsx(path, base_name: str, units: List[Tuple]) -> List[Tuple]:
    """多发票 xlsx → 物理拆成一文件一发票（每单元一个 xlsx，保留样式），落盘 uploads。

    返回 [(落盘路径, 文件名, 文件哈希), ...]；哈希由 原文件哈希+单元 确定性派生（重处理 UPSERT 去重）。
    """
    import openpyxl
    from core import config, storage
    src_wb = openpyxl.load_workbook(str(path), data_only=True)
    stem = Path(base_name).stem
    src_hash = storage.sha256_of_file(path)
    n = len(units)
    out: List[Tuple] = []
    for idx, (sheet, r0, r1) in enumerate(units, start=1):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = (sheet or "Invoice")[:31]
        _copy_range(src_wb[sheet], ws, r0, r1)
        sub_hash = storage.sha256_of_bytes(f"{src_hash}:{sheet}:{r0}-{r1}".encode())
        sub_name = f"{stem}_发票{idx}of{n}.xlsx"
        dest = config.UPLOAD_DIR / f"{sub_hash[:12]}_{sub_name}"
        wb.save(dest)
        out.append((dest, sub_name, sub_hash))
    return out


# ---- 自渲染（保真显示）------------------------------------------------

_FONT_CACHE: dict = {}


def _font(bold: bool, size_px: int):
    from PIL import ImageFont
    key = (bold, size_px)
    if key not in _FONT_CACHE:
        name = "DejaVuSans-Bold.ttf" if bold else "DejaVuSans.ttf"
        font = None
        for d in _FONT_DIRS:                         # 随包字体优先 → 系统字体兜底
            try:
                font = ImageFont.truetype(f"{d}/{name}", size_px)
                break
            except Exception:
                continue
        _FONT_CACHE[key] = font or ImageFont.load_default()   # 都没有才退位图字体
    return _FONT_CACHE[key]


def render_png(path, scale: float = 2.0) -> bytes:
    """按合成坐标自渲染 xlsx 为 PNG（加粗/字色/填充 + 长文本换行），与 words_geom 对齐。"""
    from PIL import Image, ImageDraw
    import io
    placed, page_w, page_h = _place(path)
    W, H = max(int(page_w * scale), 1), max(int(page_h * scale), 1)
    img = Image.new("RGB", (W, H), "white")
    draw = ImageDraw.Draw(img)
    for c in placed:
        x0, y0 = c["x0"] * scale, c["y0"] * scale
        if c["fill"]:
            draw.rectangle([x0, y0, x0 + _COLW * scale, c["y1"] * scale], fill=c["fill"])
        font = _font(c["bold"], max(8, int(c["size"] * scale)))
        for i, ln in enumerate(c["lines"]):
            draw.text((x0 + 2 * scale, (c["y0"] + i * _LINEH) * scale + 2),
                      ln, font=font, fill=c["color"] or (0, 0, 0))
    buf = io.BytesIO()
    img.save(buf, "PNG")
    return buf.getvalue()
