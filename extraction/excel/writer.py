"""Excel 输出（openpyxl）：八个 Sheet 的财务复核工作底稿。

Sheet 顺序：Invoice Summary / Line Items / Payment Details / Bank Details /
Validation Issues / Raw Text Archive / File Audit Trail / Change Log。
金额用数字格式；低置信度/高风险/异常/未 approve 字段高亮；冻结表头。
人工审核相关列保留占位（本期不做交互）。
"""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from typing import Any, List, Optional

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from core.models import Invoice

# ---- 样式 ----------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")    # 浅黄：warning
ERROR_FILL = PatternFill("solid", fgColor="F8CBAD")   # 浅橙红：error
CRITICAL_FILL = PatternFill("solid", fgColor="FF7C80")  # 红：critical/高风险
MONEY_FMT = "#,##0.00"


def _val(inv: Invoice, key: str) -> Any:
    return inv.f(key).value


def _money_format(value: Decimal) -> str:
    """按真实小数位生成数字格式：≤2 位用两位，>2 位显示实际精度，避免视觉截断。"""
    dp = max(0, -value.as_tuple().exponent)
    if dp <= 2:
        return MONEY_FMT
    return "#,##0." + "0" * dp


def _money_cell(ws, row: int, col: int, inv: Invoice, key: str) -> None:
    """写金额单元格：有标准化值则数字格式（保留真实精度）；解析失败则写原始文本并红色。"""
    fv = inv.f(key)
    cell = ws.cell(row=row, column=col)
    if isinstance(fv.value, Decimal):
        cell.value = float(fv.value)
        cell.number_format = _money_format(fv.value)
        if fv.suspicious:
            cell.fill = WARN_FILL
            cell.comment = Comment(f"原始文本: {fv.raw}\n{fv.note}", "system")
    elif fv.raw:
        cell.value = fv.raw
        cell.fill = ERROR_FILL
        cell.comment = Comment(f"金额未通过解析，保留原始文本: {fv.raw}", "system")
    else:
        cell.value = None


def _write_header(ws, headers: List[str]) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _autosize(ws, max_w: int = 48) -> None:
    for col in ws.columns:
        length = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is not None:
                length = max(length, min(len(str(v)), max_w))
        ws.column_dimensions[letter].width = max(10, length + 2)


# ---- Sheet 1: Invoice Summary -------------------------------------------
def _period(inv: Invoice) -> Optional[str]:
    s = _val(inv, "service_start")
    e = _val(inv, "service_end")
    if s or e:
        return f"{s or '?'} ~ {e or '?'}"
    return None


def _field(key: str):
    return lambda inv: _val(inv, key)


def _payment_methods(inv: Invoice) -> str:
    """逐个列出付款去向「方式（链）: 地址」，多个去向换行分隔，便于直接核对。"""
    parts = [f"{label}: {addr}" if addr else label
             for label, addr in inv.distinct_payment_targets()]
    return "\n".join(parts)


def _payment_comment(inv: Invoice) -> str:
    n = len(inv.distinct_payment_targets())
    return f"检测到 {n} 个付款方式/收款去向，付款前请重点核对应支付给哪一方。"


# 列规格：(表头, 类型, 取值)。类型 money 的取值为字段键，其余为取值函数。
_SUMMARY_COLS = [
    ("文件名", "text", lambda inv: inv.file_name),
    ("文件哈希", "text", lambda inv: inv.file_hash),
    ("发票号", "text", _field("invoice_no")),
    ("发票日期", "text", _field("invoice_date")),
    ("付款到期日", "text", _field("payment_due_date")),
    ("服务期间", "text", _period),
    ("Fund Valuation Date", "text", _field("fund_valuation_date")),
    ("Invoice Ccy", "text", _field("invoice_ccy_raw")),
    ("显示金额符号", "text", _field("currency_display_symbol")),
    ("结算币种", "text", _field("currency_settlement")),
    ("供应商(开票方)", "text", _field("issuer_name")),
    ("开票方地址", "text", _field("issuer_address")),
    ("开票方邮箱", "text", _field("issuer_email")),
    ("开票方电话", "text", _field("issuer_phone")),
    ("客户(收票方)", "text", _field("customer_name")),
    ("客户地址", "text", _field("customer_address")),
    ("客户邮箱", "text", _field("contact_email")),
    ("客户电话", "text", _field("contact_phone")),
    ("税前金额", "money", "subtotal"),
    ("税额", "money", "sales_tax"),
    ("总金额", "money", "total_due"),
    ("付款方式", "payment", _payment_methods),   # 列内逐行展开「方式: 地址」
    ("建议分类", "text", lambda inv: inv.classification.category),
    ("建议会计科目", "text", lambda inv: inv.classification.account),
    ("整体OCR质量等级", "text", lambda inv: inv.ocr_quality_level),
    ("整体OCR置信度", "text", lambda inv: round(inv.ocr_quality, 4)),
    ("关键字段置信度", "text", lambda inv: round(inv.key_field_confidence, 4)),
    ("金额字段置信度", "text", lambda inv: round(inv.amount_field_confidence, 4)),
    ("小数点字符置信度", "text", lambda inv: round(inv.decimal_confidence, 4)),
    ("风险评分", "risk", lambda inv: inv.risk_score),
    ("二次识别次数", "text", lambda inv: inv.recheck_count),
    ("是否重点人工审核", "text", lambda inv: "是" if inv.needs_manual_review else "否"),
    ("是否Critical Review", "text", lambda inv: "是" if inv.critical_review else "否"),
    ("校验状态", "valstatus", lambda inv: inv.validation_status),
    ("复核状态", "text", lambda inv: inv.review_status),
    ("approve状态", "approve", lambda inv: inv.approve_status),
    ("人工纠错状态", "text", lambda inv: inv.correction_status or ""),
    ("学习规则状态", "text", lambda inv: inv.learning_status or ""),
]


def _summary_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("Invoice Summary")
    _write_header(ws, [c[0] for c in _SUMMARY_COLS])
    for inv in invoices:
        r = ws.max_row + 1
        for ci, (_header, kind, payload) in enumerate(_SUMMARY_COLS, start=1):
            if kind == "money":
                _money_cell(ws, r, ci, inv, payload)
                continue
            cell = ws.cell(r, ci, payload(inv))
            if kind == "risk":
                if inv.risk_score > 30:
                    cell.fill = CRITICAL_FILL
                elif inv.risk_score > 0:
                    cell.fill = WARN_FILL
            elif kind == "payment":
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if inv.has_multiple_payment_methods:   # 多付款方式 -> 红色重点审核 + 批注
                    cell.fill = CRITICAL_FILL
                    cell.comment = Comment(_payment_comment(inv), "system")
            elif kind == "valstatus" and inv.validation_status == "has_issues":
                cell.fill = ERROR_FILL
            elif kind == "approve" and inv.approve_status != "Approved":
                cell.fill = WARN_FILL
    _autosize(ws)
    for header in ("开票方地址", "客户地址"):   # 地址列适当加宽（按列名定位，免受插列影响）
        idx = next(i for i, c in enumerate(_SUMMARY_COLS, start=1) if c[0] == header)
        ws.column_dimensions[get_column_letter(idx)].width = 40
    _pay_idx = next(i for i, c in enumerate(_SUMMARY_COLS, start=1) if c[0] == "付款方式")
    ws.column_dimensions[get_column_letter(_pay_idx)].width = 44  # 付款方式含地址，固定宽度避免过宽


# ---- Sheet 2: Line Items -------------------------------------------------
def _line_items_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("Line Items")
    headers = ["文件名", "发票号", "项目编号", "描述", "服务期间", "数量", "单价",
               "税率", "行金额", "行级置信度", "备注", "来源文件"]
    _write_header(ws, headers)
    for inv in invoices:
        ino = _val(inv, "invoice_no")
        for li in inv.line_items:
            r = ws.max_row + 1
            ws.cell(r, 1, inv.file_name)
            ws.cell(r, 2, ino)
            ws.cell(r, 3, li.item_no)
            ws.cell(r, 4, li.description)
            ws.cell(r, 5, li.service_period)
            ws.cell(r, 6, float(li.quantity) if isinstance(li.quantity, Decimal) else li.quantity)
            ws.cell(r, 7, float(li.unit_price) if isinstance(li.unit_price, Decimal) else li.unit_price)
            ws.cell(r, 8, li.tax_rate)
            amt_cell = ws.cell(r, 9)
            if isinstance(li.amount, Decimal):
                amt_cell.value = float(li.amount)
                amt_cell.number_format = _money_format(li.amount)
            elif li.amount_raw:
                amt_cell.value = li.amount_raw
                amt_cell.fill = ERROR_FILL
            elif li.note:
                amt_cell.fill = WARN_FILL
            ws.cell(r, 10, round(li.line_confidence, 4))
            note_cell = ws.cell(r, 11, li.note)
            note_cell.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 12, li.source_file)
    _autosize(ws)
    ws.column_dimensions["D"].width = 50


# ---- Sheet 3: Payment Details -------------------------------------------
def _payment_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("Payment Details")
    headers = ["文件名", "发票号", "付款方式", "付款链", "钱包地址/账户", "结算币种",
               "付款状态", "地址格式有效", "备注", "原始文本", "来源文件"]
    _write_header(ws, headers)
    for inv in invoices:
        ino = _val(inv, "invoice_no")
        for p in inv.payments:
            r = ws.max_row + 1
            ws.cell(r, 1, inv.file_name)
            ws.cell(r, 2, ino)
            ws.cell(r, 3, p.method)
            ws.cell(r, 4, p.chain)
            addr = ws.cell(r, 5, p.wallet_address)
            ws.cell(r, 6, p.settlement_currency)
            ws.cell(r, 7, p.payment_status)
            ws.cell(r, 8, "是" if p.valid_address else "否")
            ws.cell(r, 9, p.note)
            raw = ws.cell(r, 10, p.raw)
            raw.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(r, 11, p.source_file)
            if not p.valid_address:
                addr.fill = ERROR_FILL
            if p.method and p.method.startswith("Other"):
                ws.cell(r, 3).fill = WARN_FILL
    _autosize(ws)
    ws.column_dimensions["E"].width = 46
    ws.column_dimensions["J"].width = 50


# ---- Sheet 4: Bank Details / Sheet 5: Validation Issues -----------------
_SEV_FILL = {"warning": WARN_FILL, "error": ERROR_FILL, "critical": CRITICAL_FILL}


def _bank_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    """银行收款明细（结构化）：与加密钱包(Payment Details)并列的法币收款去向。"""
    ws = wb.create_sheet("Bank Details")
    _bank_keys = ("bank_name", "bank_account_name", "bank_account_no", "bank_swift")
    _write_header(ws, ["文件名", "发票号", "开户行", "户名", "账号/IBAN", "SWIFT/BIC"])
    for inv in invoices:
        if not any(inv.f(k).value for k in _bank_keys):
            continue
        r = ws.max_row + 1
        ws.cell(r, 1, inv.file_name)
        ws.cell(r, 2, _val(inv, "invoice_no"))
        ws.cell(r, 3, _val(inv, "bank_name"))
        ws.cell(r, 4, _val(inv, "bank_account_name"))
        ws.cell(r, 5, _val(inv, "bank_account_no"))
        ws.cell(r, 6, _val(inv, "bank_swift"))
    _autosize(ws)


def _issues_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("Validation Issues")
    headers = ["文件名", "发票号", "严重级别", "问题代码", "字段", "说明"]
    _write_header(ws, headers)
    for inv in invoices:
        ino = _val(inv, "invoice_no")
        for iss in inv.issues:
            r = ws.max_row + 1
            ws.cell(r, 1, inv.file_name)
            ws.cell(r, 2, ino)
            sev = ws.cell(r, 3, iss.severity)
            ws.cell(r, 4, iss.code)
            ws.cell(r, 5, iss.field)
            ws.cell(r, 6, iss.message)
            fill = _SEV_FILL.get(iss.severity)
            if fill:
                sev.fill = fill
    _autosize(ws)


# ---- Sheet 6: Raw Text Archive ------------------------------------------
def _raw_text_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("Raw Text Archive")
    headers = ["文件名", "文件哈希", "解析方式", "完整PDF文本", "完整OCR原文"]
    _write_header(ws, headers)
    for inv in invoices:
        r = ws.max_row + 1
        ws.cell(r, 1, inv.file_name)
        ws.cell(r, 2, inv.file_hash)
        ws.cell(r, 3, inv.parse_method)
        c4 = ws.cell(r, 4, inv.raw_pdf_text)
        c5 = ws.cell(r, 5, inv.raw_ocr_text)
        c4.alignment = Alignment(wrap_text=True, vertical="top")
        c5.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 70
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 24


# ---- Sheet 7: File Audit Trail ------------------------------------------
def _audit_sheet(wb: Workbook, invoices: List[Invoice]) -> None:
    ws = wb.create_sheet("File Audit Trail")
    headers = ["文件名", "上传时间", "文件哈希", "解析方式", "是否OCR", "OCR引擎",
               "首次处理时间", "二次识别时间", "二次识别原因", "风险评分", "处理状态",
               "人工审核状态", "approve人", "approve时间"]
    _write_header(ws, headers)
    for inv in invoices:
        r = ws.max_row + 1
        ws.cell(r, 1, inv.file_name)
        ws.cell(r, 2, inv.uploaded_at)
        ws.cell(r, 3, inv.file_hash)
        ws.cell(r, 4, inv.parse_method)
        ws.cell(r, 5, "是" if inv.ocr_used else "否")
        ws.cell(r, 6, inv.ocr_engine)
        ws.cell(r, 7, inv.processed_at)
        ws.cell(r, 8, "")       # 二次识别时间（占位）
        ws.cell(r, 9, "")       # 二次识别原因（占位）
        ws.cell(r, 10, inv.risk_score)
        ws.cell(r, 11, inv.parse_status)
        ws.cell(r, 12, inv.review_status)
        ws.cell(r, 13, "")      # approve 人（人工审核阶段填）
        ws.cell(r, 14, "")      # approve 时间
    _autosize(ws)


# ---- Sheet 8: Change Log（占位，人工审核阶段写入）-----------------------
def _change_log_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Change Log")
    headers = ["字段名", "修改前值", "修改后值", "修改人", "修改时间", "修改原因",
               "来源文件", "来源页面", "来源区域", "是否用于学习", "生成规则类型", "规则状态"]
    _write_header(ws, headers)
    note = ws.cell(2, 1, "（人工审核与纠错为后续阶段，本表暂为空）")
    note.font = Font(italic=True, color="888888")
    _autosize(ws)


def build_workbook(invoices: List[Invoice], out_path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认空 Sheet
    _summary_sheet(wb, invoices)
    _line_items_sheet(wb, invoices)
    _payment_sheet(wb, invoices)
    _bank_sheet(wb, invoices)
    _issues_sheet(wb, invoices)
    _raw_text_sheet(wb, invoices)
    _audit_sheet(wb, invoices)
    _change_log_sheet(wb)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
