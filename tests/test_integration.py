"""端到端集成：一个跨境小主体的完整账套，串起所有增量并验证三表勾稽全过。

建账期初 → 外币采购+本币销售发票入账 → 外币结算(已实现汇兑损益)+本币结算 →
非发票流水入账(手续费/利息) → 折旧 → 期末外币敞口重估报告 → 期末结转 →
三张 IFRS 报表勾稽 E1/E2/E3/E4/E6 全过。隔离临时库与临时汇率文件，绝不碰真实数据。
"""
import json
import shutil
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

import fitz
from openpyxl import load_workbook

from core import config, db, fx
from core.models import Classification, FieldValue, Invoice, Transaction
from extraction import pipeline
from ledger import accounts as A
from ledger import service as ledger
from reports import service as reports
from review import service as review

D = Decimal


def _invoice_pdf_bytes():
    """生成一份确定性的「原始发票文件」。

    这里不直接构造 Invoice 对象：验收必须从真实上传入口起步，同时避免把
    客户原件或未进公开包的 `发票/` 目录变成测试依赖。
    """
    doc = fitz.open()
    page = doc.new_page()
    lines = [
        "Acme Consulting LLC",
        "Bill To: Acceptance Test Company",
        "Invoice No: AC-2026-9001",
        "Invoice Date: 2026-06-15",
        "Currency: USD",
        "Subtotal: $1,000.00",
        "Tax: $100.00",
        "Total Due: $1,100.00",
    ]
    for i, line in enumerate(lines):
        page.insert_text((40, 50 + i * 22), line)
    data = doc.tobytes()
    doc.close()
    return data


def _inv(no, total, fh, ccy="USD", account="6440 会计费 Accounting Fees",
         issuer="Vendor", customer="MyCo", date="2026-06-05"):
    inv = Invoice()
    inv.file_hash = fh; inv.doc_type = "invoice"; inv.approve_status = "Approved"
    inv.set("invoice_no", FieldValue(value=no))
    inv.set("invoice_date", FieldValue(value=date))
    inv.set("total_due", FieldValue(value=total))
    inv.set("currency_settlement", FieldValue(value=ccy))
    inv.set("issuer_name", FieldValue(value=issuer))
    inv.set("customer_name", FieldValue(value=customer))
    inv.classification = Classification(account=account)
    db.save_invoice(inv)
    return inv


class CrossBorderEndToEndTest(unittest.TestCase):
    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())
        self._db, self._up, self._fx = config.DB_PATH, config.UPLOAD_DIR, config.FX_RATES_PATH
        self._af = config.FX_AUTO_FETCH
        config.DB_PATH = self._dir / "t.db"; config.UPLOAD_DIR = self._dir / "up"
        config.FX_RATES_PATH = self._dir / "fx.json"
        config.FX_AUTO_FETCH = False        # 纯本地、不联网（测试隔离）
        config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        config.FX_RATES_PATH.write_text(json.dumps({"EUR": [
            {"date": "2026-06-01", "rate": "1.10"},   # 录入日(as_of=2026-06-05)取此条
            {"date": "2026-06-15", "rate": "1.15"},   # 结算日
            {"date": "2026-06-30", "rate": "1.20"}]}), encoding="utf-8")   # 期末
        db._initialized = False; db.init_db()

    def tearDown(self):
        config.DB_PATH, config.UPLOAD_DIR, config.FX_RATES_PATH = self._db, self._up, self._fx
        config.FX_AUTO_FETCH = self._af
        db._initialized = False
        shutil.rmtree(self._dir, ignore_errors=True)

    def test_full_cross_border_cycle_ties_out(self):
        # ① 建账期初：银行 50000 + 固定资产 12000 / 实收资本 62000
        ledger.post_opening(other_lines=[
            {"account": "1002 银行存款 Bank", "amount": "50000", "side": "debit"},
            {"account": "1500 固定资产 Fixed Assets", "amount": "12000", "side": "debit"},
            {"account": "3000 实收资本 Share Capital", "amount": "62000", "side": "credit"}],
            date="2026-05-31")

        # ② 发票入账：外币采购 EUR1000@1.10=1100(AP1)、本币销售 USD3000(AR)、外币采购 EUR500@1.10=550(AP2,不结算)
        _inv("EU-P1", "1000.00", "ap1", ccy="EUR")
        ledger.post_invoice_by_hash("ap1", direction="AP", as_of="2026-06-05")   # 录入日汇率 1.10
        _inv("US-S1", "3000.00", "ar1", ccy="USD", account="4000 主营业务收入 Revenue")
        ledger.post_invoice_by_hash("ar1", direction="AR")
        _inv("EU-P2", "500.00", "ap2", ccy="EUR")
        ledger.post_invoice_by_hash("ap2", direction="AP", as_of="2026-06-05")   # 录入日汇率 1.10

        # ③ 结算：外币 AP1 结算日1.15 付 EUR1000=1150 → 汇兑损失 50；本币 AR 收 3000
        ledger.settle_invoice("ap1", cash_amount="1000", cash_currency="EUR",
                              settle_amount="1100", date="2026-06-15", diff_reason="fx_gain_loss")
        ledger.settle_invoice("ar1", cash_amount="3000", date="2026-06-16")

        # ④ 非发票流水入账：手续费 20(支出)、利息 10(收入)
        st = Invoice(); st.file_hash = "st1"; st.doc_type = "statement"
        st.set("currency_settlement", FieldValue(value="USD"))
        st.transactions = [Transaction(date="2026-06-20", description="Bank fee", expense=D("20")),
                           Transaction(date="2026-06-21", description="Interest", income=D("10"))]
        db.save_invoice(st)
        ledger.post_statement_entry("st1", 0, A.FEE, activity="operating")
        ledger.post_statement_entry("st1", 1, "4100 利息收入 Interest Income", activity="operating")

        # ⑤ 折旧 200（非现金）
        ledger.post_manual_entry([{"account": "6602 折旧摊销费 Depreciation & Amortization", "debit": "200"},
                                  {"account": "1509 累计折旧 Accumulated Depreciation", "credit": "200"}],
                                 date="2026-06-25")

        # ⑥ 期末外币敞口重估报告（AP2 未结 EUR500 @期末1.20=600，账面550 → 未实现损失 50）
        rev = ledger.fx_revaluation_view("2026-06-30")
        eur = [c for c in rev["by_currency"] if c["currency"] == "EUR"][0]
        self.assertEqual(D(eur["open_ccy"]), D("500.00"))
        self.assertEqual(D(eur["book_usd"]), D("550.00"))
        self.assertEqual(D(eur["reval_usd"]), D("600.00"))
        self.assertEqual(D(rev["total_unrealized_pl"]), D("-50.00"))   # AP 升值→未实现损失

        # 结算状态：AP1/AR 已清零、AP2 未结 550
        opens = {o["file_hash"]: D(o["open"]) for o in ledger.open_view()}
        self.assertNotIn("ap1", opens); self.assertNotIn("ar1", opens)
        self.assertEqual(opens.get("ap2"), D("550.00"))

        # ⑦ 关账前勾稽全过（含 E4 直接法=间接法）
        ck = reports.checks()
        for k in ("E1_balance_sheet_balanced", "E2_income_to_retained", "E3_cash_tie",
                  "E4_cfo_direct_indirect", "E6_all_classified"):
            self.assertTrue(ck[k]["ok"], f"{k} 未过：{ck[k]}")
        self.assertTrue(ck["can_issue"])

        # 直接法经营现金流 = 收3000 − 付AP1 1150 − 手续费20 + 利息10 = 1840
        cf = reports.cash_flow_statement()
        op = [l for l in cf["lines"] if l["key"] == A.OPERATING][0]
        self.assertEqual(D(op["amount"]), D("1840"))
        self.assertEqual(D(reports.cash_flow_indirect()["operating"]), D("1840"))

        # ⑧ 期末结转后仍全过（利润表排除结转、E2/E4 恒等）
        ledger.close_period("2026-06")
        ck2 = reports.checks()
        self.assertTrue(ck2["can_issue"], f"关账后勾稽应仍通过：{ck2}")
        led = ledger.load_ledger()
        dr, cr, _ = led.trial_balance()
        self.assertEqual(dr, cr)


class RawFileToReportAcceptanceTest(unittest.TestCase):
    """单条 MVP 验收证据：原始 PDF → 提取/入库 → 人工审核 → 过账/关账 → 报表 Excel。"""

    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())
        self._paths = {name: getattr(config, name) for name in
                       ("DB_PATH", "UPLOAD_DIR", "EXPORT_DIR")}
        self._auto_fetch = config.FX_AUTO_FETCH
        config.DB_PATH = self._dir / "acceptance.db"
        config.UPLOAD_DIR = self._dir / "uploads"
        config.EXPORT_DIR = self._dir / "exports"
        config.FX_AUTO_FETCH = False              # 验收全程离线，且本例为 USD
        config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        config.EXPORT_DIR.mkdir(parents=True, exist_ok=True)
        db._initialized = False
        db.init_db()

    def tearDown(self):
        for name, value in self._paths.items():
            setattr(config, name, value)
        config.FX_AUTO_FETCH = self._auto_fetch
        db._initialized = False
        shutil.rmtree(self._dir, ignore_errors=True)

    def test_raw_invoice_to_auditable_financial_statements_excel(self):
        # 1) 从正式上传入口读真实 PDF 字节：原件落盘，发票字段提取并入库。
        raw = _invoice_pdf_bytes()
        extracted = pipeline.process_upload(raw, "acceptance_invoice.pdf", "invoice")
        self.assertEqual(len(extracted), 1)
        inv = extracted[0]
        self.assertEqual(inv.parse_status, "parsed")
        self.assertEqual(inv.f("invoice_no").value, "AC-2026-9001")
        self.assertEqual(inv.f("invoice_date").value, "2026-06-15")
        self.assertEqual(D(inv.f("total_due").value), D("1100.00"))
        self.assertEqual(Path(inv.file_path).read_bytes(), raw)

        # 2) 人工审核是硬闸门：未通过时禁止入账；人确认分类并 Approve 后才放行。
        with self.assertRaises(ValueError):
            ledger.post_invoice_by_hash(inv.file_hash, direction="AP", by="acceptance")
        review.set_classification(inv.file_hash, "Professional Service",
                                  "6440 会计费 Accounting Fees",
                                  changed_by="acceptance", reason="MVP 自动化验收")
        approved = review.act(inv.file_hash, "Approved", by="acceptance")
        self.assertEqual(approved["approve_status"], "Approved")

        # 3) 应计过账、试算平衡、期末关账，关账后 E1/E2/E3/E4/E6 仍须全过。
        entry_no = ledger.post_invoice_by_hash(inv.file_hash, direction="AP", by="acceptance")
        self.assertTrue(entry_no.startswith("202606-"))
        _dr, _cr, _rows, balanced = ledger.trial_balance()
        self.assertTrue(balanced)
        ledger.close_period("2026-06", by="acceptance")
        checks = reports.checks()
        self.assertTrue(checks["can_issue"], checks)
        for key in ("E1_balance_sheet_balanced", "E2_income_to_retained", "E3_cash_tie",
                    "E4_cfo_direct_indirect", "E6_all_classified"):
            self.assertTrue(checks[key]["ok"], f"{key} 未通过：{checks[key]}")

        # 4) 从正式报表出口导出 Excel，并验证七张表和「发票→分录」审计轨迹。
        out = reports.export_excel("acceptance_financial_statements.xlsx")
        self.assertTrue(out.is_file())
        workbook = load_workbook(out, data_only=False)
        try:
            expected_sheets = {
                "封面 Cover", "利润表 Income Statement", "资产负债表 Balance Sheet",
                "现金流量表 Cash Flow", "勾稽校验 Reconciliation",
                "科目余额 取数轨迹", "分录明细 凭证轨迹",
            }
            self.assertEqual(set(workbook.sheetnames), expected_sheets)
            audit_values = {
                str(cell.value)
                for row in workbook["分录明细 凭证轨迹"].iter_rows()
                for cell in row if cell.value is not None
            }
            self.assertTrue(any("AC-2026-9001" in value for value in audit_values), audit_values)
        finally:
            workbook.close()                         # Windows 下也释放 xlsx 文件句柄


if __name__ == "__main__":
    unittest.main()
