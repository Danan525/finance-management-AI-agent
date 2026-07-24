"""Excel 导出(writer.build_workbook)输出验证——此前核心交付物无回归网(自检补)。"""
import shutil
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path


def _has_openpyxl():
    try:
        import openpyxl  # noqa
        return True
    except Exception:
        return False


@unittest.skipUnless(_has_openpyxl(), "需要 openpyxl")
class ExcelWriterTest(unittest.TestCase):
    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())

    def tearDown(self):
        shutil.rmtree(self._dir, ignore_errors=True)

    def _inv(self):
        from core.models import Invoice, FieldValue, LineItem, ValidationIssue
        inv = Invoice(file_hash="h1", file_name="a.pdf", doc_type="invoice",
                      uploaded_at="2026-07-22T00:00:00", processed_at="2026-07-22T00:00:00",
                      parse_status="ok", review_status="Pending")
        inv.set("invoice_no", FieldValue(raw="INV-1", value="INV-1"))
        inv.set("subtotal", FieldValue(raw="1,000.00", value=Decimal("1000.00")))
        inv.set("sales_tax", FieldValue(raw="234.56", value=Decimal("234.56")))
        inv.set("total_due", FieldValue(raw="1,234.56", value=Decimal("1234.56")))
        inv.set("bank_name", FieldValue(raw="Test Bank", value="Test Bank"))
        inv.line_items = [LineItem(description="Consulting", amount=Decimal("1000.00"),
                                   amount_raw="1,000.00", source_file="a.pdf")]
        inv.issues = [ValidationIssue("TOTAL_MISMATCH", "示例不平", "total_due", "error")]
        return inv

    def _cells(self, ws):
        return [c.value for row in ws.iter_rows() for c in row]

    def test_workbook_structure_and_values(self):
        from extraction.excel import writer
        import openpyxl
        out = writer.build_workbook([self._inv()], self._dir / "out.xlsx")
        self.assertTrue(out.exists())
        wb = openpyxl.load_workbook(out)
        # 8 个 Sheet、顺序固定
        self.assertEqual(wb.sheetnames, [
            "Invoice Summary", "Line Items", "Payment Details", "Bank Details",
            "Validation Issues", "Raw Text Archive", "File Audit Trail", "Change Log"])
        # 合计值以数值写入 + 两位小数格式
        summ = wb["Invoice Summary"]
        money = [c for row in summ.iter_rows() for c in row
                 if isinstance(c.value, (int, float)) and abs(c.value - 1234.56) < 1e-6]
        self.assertTrue(money, "Invoice Summary 应含数值 1234.56")
        self.assertIn(".00", money[0].number_format)          # 声明两位小数格式
        # 明细金额、校验 issue 均落表
        self.assertIn(1000.0, [c.value for row in wb["Line Items"].iter_rows() for c in row])
        self.assertTrue(any("TOTAL_MISMATCH" in str(v) for v in self._cells(wb["Validation Issues"])))
        self.assertTrue(any("Test Bank" in str(v) for v in self._cells(wb["Bank Details"])))

    def test_empty_invoice_list_ok(self):
        """空列表也应产出结构完整的工作簿、不崩。"""
        from extraction.excel import writer
        import openpyxl
        out = writer.build_workbook([], self._dir / "empty.xlsx")
        wb = openpyxl.load_workbook(out)
        self.assertEqual(len(wb.sheetnames), 8)


if __name__ == "__main__":
    unittest.main()
