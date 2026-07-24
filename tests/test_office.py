"""旧版办公格式（.doc 等）经 LibreOffice 转文本 PDF 后自动提取。"""
import shutil
import subprocess
import tempfile
import unittest
from pathlib import Path

from core import config, db
from extraction.extract import office


def _has_fitz():
    try:
        import fitz  # noqa
        return True
    except Exception:
        return False


@unittest.skipUnless(office.available() and _has_fitz(), "需 LibreOffice + PyMuPDF")
class OfficeConvertTest(unittest.TestCase):
    def setUp(self):
        self._dir = tempfile.mkdtemp()
        self._db, self._up = config.DB_PATH, config.UPLOAD_DIR
        config.DB_PATH = Path(self._dir) / "t.db"
        config.UPLOAD_DIR = Path(self._dir) / "up"
        config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        db._initialized = False
        db.init_db()

    def tearDown(self):
        config.DB_PATH, config.UPLOAD_DIR = self._db, self._up
        db._initialized = False
        shutil.rmtree(self._dir, ignore_errors=True)

    def _make_doc(self, text: str) -> Path:
        """用 LibreOffice 从 txt 造一个 .doc 样本。"""
        txt = Path(self._dir) / "src.txt"
        txt.write_text(text, encoding="utf-8")
        with tempfile.TemporaryDirectory() as prof:
            subprocess.run([office.soffice_bin(), "--headless", "--norestore",
                            f"-env:UserInstallation=file://{prof}",
                            "--convert-to", "doc", "--outdir", self._dir, str(txt)],
                           check=True, timeout=120, capture_output=True)
        return Path(self._dir) / "src.doc"

    def test_doc_converted_to_text_pdf_and_extracted(self):
        from extraction import pipeline
        doc = self._make_doc(
            "INVOICE\nNorthbridge Labs LLC\n1400 Market Test Ave, Suite 210, Seattle, WA 98101\n"
            "Invoice No.: DOC-77\nInvoice Date: 2026-07-01\n"
            "Consulting and advisory services for the billing period ending 2026-07-01\n"
            "Total Due: USD 55.00\n")
        inv = pipeline.process_local(doc)[0]
        self.assertEqual(inv.parse_method, "pdf_text")        # 转 .docx 后走文本路径（fitz 原生读）
        self.assertNotEqual(inv.parse_status, "failed")       # 不再是失败记录
        self.assertEqual(inv.f("invoice_no").value, "DOC-77")
        self.assertTrue(inv.page_sizes)                        # 可渲染预览
        # 文书类改转 .docx（表格用 fitz 原生逐格读，避免转 PDF 时单元格被粘连成一行）
        self.assertTrue(str(inv.file_path).endswith(".docx"))

    def test_xls_converted_to_xlsx_and_extracted(self):
        """电子表格（.xls）改转 .xlsx 走 openpyxl 读全量单元格（转 PDF 会按列宽裁剪文字）。"""
        import openpyxl
        from extraction import pipeline
        src = Path(self._dir) / "s.xlsx"
        wb = openpyxl.Workbook(); ws = wb.active
        ws["A1"] = "Meridian Freight Co"; ws["A2"] = "Invoice No.: XLS-88"
        ws["A3"] = "Invoice Date: 2026-07-02"
        ws["A5"] = "Description"; ws["B5"] = "Amount"
        ws["A6"] = "Ocean freight"; ws["B6"] = 1200.00
        ws["A7"] = "Total Due"; ws["B7"] = 1200.00
        wb.save(str(src))
        with tempfile.TemporaryDirectory() as prof:
            subprocess.run([office.soffice_bin(), "--headless", "--norestore",
                            f"-env:UserInstallation=file://{prof}",
                            "--convert-to", "xls", "--outdir", self._dir, str(src)],
                           check=True, timeout=120, capture_output=True)
        inv = pipeline.process_local(str(Path(self._dir) / "s.xls"))[0]
        self.assertEqual(inv.parse_method, "excel")            # 转 .xlsx 后走 Excel 路径
        self.assertNotEqual(inv.parse_status, "failed")
        self.assertEqual(inv.f("invoice_no").value, "XLS-88")
        self.assertTrue(str(inv.file_path).endswith(".xlsx"))

    def test_is_convertible(self):
        self.assertTrue(office.is_convertible(".doc"))
        self.assertTrue(office.is_convertible(".xls"))
        self.assertFalse(office.is_convertible(".docx"))       # OOXML 走专用路径
        self.assertFalse(office.is_convertible(".pdf"))


if __name__ == "__main__":
    unittest.main()
