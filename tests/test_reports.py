"""报表中心：利润表/资产负债表取数 + 勾稽（E1 资产=负债+权益、E6 归类完整、勾稽不过不出表）。

隔离临时库。报表从 module 6 已过账分录取数。"""
import shutil
import tempfile
import unittest
from decimal import Decimal
from pathlib import Path

from core import config, db
from core.models import Classification, FieldValue, Invoice
from ledger import service as ledger
from ledger import store
from ledger.engine import JournalEntry, JournalLine
from reports import service as reports

D = Decimal


def _inv(no, total, fh, account="6440 会计费 Accounting Fees"):
    inv = Invoice()
    inv.file_hash = fh
    inv.doc_type = "invoice"
    inv.approve_status = "Approved"
    inv.set("invoice_no", FieldValue(value=no))
    inv.set("invoice_date", FieldValue(value="2026-06-01"))
    inv.set("total_due", FieldValue(value=total))
    inv.set("issuer_name", FieldValue(value="Vendor"))
    inv.classification = Classification(account=account)
    db.save_invoice(inv)
    return inv


class ReportsTest(unittest.TestCase):
    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())
        self._db, self._up = config.DB_PATH, config.UPLOAD_DIR
        config.DB_PATH = self._dir / "t.db"
        config.UPLOAD_DIR = self._dir / "up"
        config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        db._initialized = False
        db.init_db()

    def tearDown(self):
        config.DB_PATH, config.UPLOAD_DIR = self._db, self._up
        db._initialized = False
        shutil.rmtree(self._dir, ignore_errors=True)

    def _opening_capital(self, amount):
        # 本期股东出资：借 银行 / 贷 实收资本（动现金 → 筹资活动流入；source_kind='manual'，
        # 非期初建账——建账另见 test_opening）。
        store.post_entry(JournalEntry(
            "2026-06-01", "股东出资", [
                JournalLine("1002 银行存款 Bank", debit=amount),
                JournalLine("3000 实收资本 Share Capital", credit=amount)],
            source_kind="manual"), by="t", at="2026-06-01T00:00:00Z", activity="financing")

    def test_basis_currency_follows_functional(self):
        # 覆盖性验证发现：报表币种标签曾硬编码 USD；非 USD 功能货币主体标签应跟随
        old = config.FUNCTIONAL_CURRENCY
        try:
            config.FUNCTIONAL_CURRENCY = "EUR"
            self.assertEqual(reports.generate()["basis"]["currency"], "EUR")
        finally:
            config.FUNCTIONAL_CURRENCY = old

    def test_empty_reports_balance(self):
        r = reports.generate()
        self.assertTrue(r["checks"]["E1_balance_sheet_balanced"]["ok"])
        self.assertTrue(r["can_issue"])

    def test_ar_sale_income_and_balance(self):
        # AR 开票 1000（Dr 应收 / Cr 收入）→ 利润表收入 1000、净利 1000；资产=负债+权益
        ledger.post_invoice(_inv("AR-1", "1000.00", "h1"), direction="AR")
        istmt = reports.income_statement()
        by = {l["key"]: l["amount"] for l in istmt["lines"]}
        self.assertEqual(by["Revenue"], "1000.00")
        self.assertEqual(by["NetIncome"], "1000.00")
        bs = reports.balance_sheet()
        self.assertTrue(bs["balanced"])
        # 净利润进入权益
        self.assertIn("CurrentNetIncome", [e["key"] for e in bs["equity"]])
        self.assertEqual(bs["diff"], "0.00")

    def test_ap_expense_reduces_net_income(self):
        self._opening_capital("5000.00")
        ledger.post_invoice(_inv("AP-1", "1000.00", "h1"))     # Dr 费用 / Cr 应付
        istmt = reports.income_statement()
        by = {l["key"]: l["amount"] for l in istmt["lines"]}
        self.assertEqual(by["Opex"], "1000.00")
        self.assertEqual(by["NetIncome"], "-1000.00")          # 纯费用 → 亏损
        bs = reports.balance_sheet()
        self.assertTrue(bs["balanced"])
        # 资产=银行5000；负债=应付1000；权益=实收5000 + 净利-1000 = 4000；5000==1000+4000
        self.assertEqual(bs["assets_total"], "5000.00")
        self.assertEqual(bs["liab_equity_total"], "5000.00")

    def test_full_cycle_accrue_settle_balances(self):
        self._opening_capital("5000.00")
        ledger.post_invoice(_inv("AP-1", "1000.00", "h1"))
        ledger.settle_invoice("h1", cash_amount="1000.00")     # 付清
        bs = reports.balance_sheet()
        self.assertTrue(bs["balanced"])
        # 银行 5000-1000=4000；应付 0；权益 5000 + 净利(-1000)=4000
        self.assertEqual(bs["assets_total"], "4000.00")
        self.assertEqual(bs["liab_equity_total"], "4000.00")

    def test_cash_flow_direct_and_e3(self):
        self._opening_capital("5000.00")                       # 筹资 +5000
        ledger.post_invoice(_inv("AP-1", "1000.00", "h1"))
        ledger.settle_invoice("h1", cash_amount="1000.00")     # 经营 -1000
        ledger.post_invoice(_inv("AR-1", "3000.00", "h2"), direction="AR")
        ledger.settle_invoice("h2", cash_amount="3000.00")     # 经营 +3000
        cf = reports.cash_flow_statement()
        by = {l["key"]: l["amount"] for l in cf["lines"]}
        self.assertEqual(by["financing"], "5000.00")
        self.assertEqual(by["operating"], "2000.00")
        self.assertEqual(by["investing"], "0")
        self.assertEqual(cf["net_change"], "7000.00")
        self.assertEqual(cf["ending"], "7000.00")
        self.assertEqual(cf["bs_cash"], "7000.00")
        self.assertTrue(cf["e3_ok"])                           # CFS 期末现金 == BS 货币资金
        self.assertTrue(reports.checks()["can_issue"])

    def test_investing_activity_inferred(self):
        # 固定资产发票 → 结算现金流归投资活动
        ledger.post_invoice(_inv("AP-FA", "2000.00", "h1",
                                 account="1500 固定资产 Fixed Assets"))
        ledger.settle_invoice("h1", cash_amount="2000.00")
        cf = reports.cash_flow_statement()
        by = {l["key"]: l["amount"] for l in cf["lines"]}
        self.assertEqual(by["investing"], "-2000.00")
        self.assertEqual(by["operating"], "0")

    def test_cash_entry_requires_activity(self):
        # 动现金却不标活动 → 拒绝（E5 前提）
        with self.assertRaises(ValueError):
            store.post_entry(JournalEntry("2026-06-01", "付现无分类", [
                JournalLine("6440 会计费 Accounting Fees", debit="100"),
                JournalLine("1002 银行存款 Bank", credit="100")],
                source_kind="manual"), by="t", at="2026-06-01T00:00:00Z")

    def test_internal_transfer_no_activity_not_in_cfs(self):
        self._opening_capital("5000.00")
        # 银行取现到库存现金：两边都是现金及等价物 → 净流 0 → 不标活动、不进 CFS
        store.post_entry(JournalEntry("2026-06-01", "取现", [
            JournalLine("1001 现金 Cash on Hand", debit="1000"),
            JournalLine("1002 银行存款 Bank", credit="1000")],
            source_kind="manual"), by="t", at="2026-06-01T00:00:00Z")
        # 误给内部腾挪标活动 → 拒绝
        with self.assertRaises(ValueError):
            store.post_entry(JournalEntry("2026-06-01", "取现误标", [
                JournalLine("1001 现金 Cash on Hand", debit="1"),
                JournalLine("1002 银行存款 Bank", credit="1")],
                source_kind="manual"), by="t", at="2026-06-01T00:00:00Z", activity="operating")
        cf = reports.cash_flow_statement()
        self.assertEqual(cf["net_change"], "5000.00")          # 只有筹资 5000，取现不计
        self.assertTrue(cf["e3_ok"])                           # 期末现金仍 = 银行+现金合计

    def test_unclassified_blocks_issue(self):
        # 造一个无法归类的科目（未知类别编码 9xxx）→ E6 失败、不出表
        store.post_entry(JournalEntry(
            "2026-06-01", "怪科目", [
                JournalLine("9999 未知科目 Weird", debit="100"),
                JournalLine("1002 银行存款 Bank", credit="100")],
            source_kind="manual"), by="t", at="2026-06-01T00:00:00Z", activity="operating")
        ck = reports.checks()
        self.assertFalse(ck["E6_all_classified"]["ok"])
        self.assertIn("9999 未知科目 Weird", ck["E6_all_classified"]["unclassified"])
        self.assertFalse(ck["can_issue"])                      # 勾稽不过不出表


    def test_expanded_chart_other_income_tax_and_contra_asset(self):
        # 其它收益(利息)加回净利、所得税减去;累计折旧抵减固定资产(PPE)
        self._opening_capital("10000.00")
        ledger.post_manual_entry([{"account": "1500 固定资产 Fixed Assets", "debit": "3000"},
                                  {"account": "1002 银行存款 Bank", "credit": "3000"}],
                                 date="2026-06-02", memo="购设备", activity="investing")
        ledger.post_manual_entry([{"account": "1002 银行存款 Bank", "debit": "50"},
                                  {"account": "4100 利息收入 Interest Income", "credit": "50"}],
                                 date="2026-06-05", memo="利息", activity="operating")
        ledger.post_manual_entry([{"account": "6602 折旧摊销费 Depreciation & Amortization", "debit": "300"},
                                  {"account": "1509 累计折旧 Accumulated Depreciation", "credit": "300"}],
                                 date="2026-06-30", memo="折旧")
        ledger.post_manual_entry([{"account": "6801 所得税费用 Income Tax Expense", "debit": "100"},
                                  {"account": "2220 应交税费-应交所得税 Income Tax Payable", "credit": "100"}],
                                 date="2026-06-30", memo="所得税")
        istmt = {l["key"]: l["amount"] for l in reports.income_statement()["lines"]}
        self.assertEqual(istmt["OtherIncome"], "50")
        self.assertEqual(istmt["IncomeTax"], "100")
        # 净利 = 其它收益50 - 折旧300 - 所得税100 = -350
        self.assertEqual(istmt["NetIncome"], "-350")
        bs = reports.balance_sheet()
        ppe = [x["amount"] for x in bs["assets"] if x["key"] == "PPE"][0]
        self.assertEqual(ppe, "2700")           # 固定资产3000 - 累计折旧300
        self.assertTrue(bs["balanced"])
        self.assertTrue(reports.checks()["can_issue"])

    def test_generate_is_json_serializable(self):
        # generate() 直接进 JSONResponse，必须无 Decimal 泄漏（曾致 /api/reports 500）
        import json
        self._opening_capital("5000.00")
        ledger.post_invoice(_inv("AP-1", "1000.00", "h1"))
        ledger.settle_invoice("h1", cash_amount="1000.00")
        json.dumps(reports.generate())        # 不抛 TypeError 即通过
        json.dumps(reports.generate())        # 空前也应可序列化

    def test_export_excel_when_balanced(self):
        from openpyxl import load_workbook
        self._opening_capital("5000.00")
        ledger.post_invoice(_inv("AP-1", "1000.00", "h1"))
        ledger.settle_invoice("h1", cash_amount="1000.00")
        out = reports.export_excel()
        self.assertTrue(out.exists())
        wb = load_workbook(str(out))
        for sheet in ("封面 Cover", "利润表 Income Statement", "资产负债表 Balance Sheet",
                      "现金流量表 Cash Flow", "勾稽校验 Reconciliation",
                      "科目余额 取数轨迹", "分录明细 凭证轨迹"):
            self.assertIn(sheet, wb.sheetnames)

    def test_export_blocked_when_unbalanced(self):
        # 未归类科目 → 勾稽不过 → 拒绝导出（勾稽不通过不出表）
        store.post_entry(JournalEntry("2026-06-01", "怪科目", [
            JournalLine("9999 未知科目 Weird", debit="100"),
            JournalLine("1002 银行存款 Bank", credit="100")],
            source_kind="manual"), by="t", at="2026-06-01T00:00:00Z", activity="operating")
        with self.assertRaises(ValueError):
            reports.export_excel()


class E4IndirectCashFlowTest(unittest.TestCase):
    """间接法现金流 + E4（直接法 == 间接法）。综合场景含应计/结算/折旧/投资/筹资。"""

    def setUp(self):
        self._dir = Path(tempfile.mkdtemp())
        self._db, self._up = config.DB_PATH, config.UPLOAD_DIR
        config.DB_PATH = self._dir / "t.db"; config.UPLOAD_DIR = self._dir / "up"
        config.UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
        db._initialized = False; db.init_db()

    def tearDown(self):
        config.DB_PATH, config.UPLOAD_DIR = self._db, self._up
        db._initialized = False
        shutil.rmtree(self._dir, ignore_errors=True)

    def _build(self):
        L = JournalLine
        # 筹资：股东注资 12000（现金+）
        ledger.post_manual_entry([{"account": "1002 银行存款 Bank", "debit": "12000"},
                                  {"account": "3000 实收资本 Share Capital", "credit": "12000"}],
                                 date="2026-06-01", activity="financing")
        # 投资：购固定资产 5000（现金−）
        ledger.post_manual_entry([{"account": "1500 固定资产 Fixed Assets", "debit": "5000"},
                                  {"account": "1002 银行存款 Bank", "credit": "5000"}],
                                 date="2026-06-02", activity="investing")
        # 应计：赊销收入 3000（不动现金，AR 控制账户）
        ledger.post_manual_entry([{"account": "1100 应收账款 Accounts Receivable", "debit": "3000"},
                                  {"account": "4000 主营业务收入 Revenue", "credit": "3000"}],
                                 date="2026-06-03", allow_control=True, counterparty="客户X")
        # 结算：收款 3000（现金+，经营）
        ledger.post_manual_entry([{"account": "1002 银行存款 Bank", "debit": "3000"},
                                  {"account": "1100 应收账款 Accounts Receivable", "credit": "3000"}],
                                 date="2026-06-04", activity="operating",
                                 allow_control=True, counterparty="客户X")
        # 经营：付手续费 500（现金−）
        ledger.post_manual_entry([{"account": "6603 财务费用-手续费 Bank/Platform Fees", "debit": "500"},
                                  {"account": "1002 银行存款 Bank", "credit": "500"}],
                                 date="2026-06-05", activity="operating")
        # 折旧 800（非现金）
        ledger.post_manual_entry([{"account": "6602 折旧摊销费 Depreciation & Amortization", "debit": "800"},
                                  {"account": "1509 累计折旧 Accumulated Depreciation", "credit": "800"}],
                                 date="2026-06-06")

    def test_e4_ties_direct_and_indirect(self):
        self._build()
        ind = reports.cash_flow_indirect()
        # 直接法经营 = 收款3000 − 手续费500 = 2500
        self.assertEqual(D(ind["direct_operating"]), D("2500"))
        # 间接法 = 净利1700 + 折旧800 − ΔAR0 = 2500
        self.assertEqual(D(ind["net_income"]), D("1700"))
        self.assertEqual(D(ind["depreciation_addback"]), D("800"))
        self.assertEqual(D(ind["operating"]), D("2500"))
        self.assertTrue(ind["e4_ok"])
        ck = reports.checks()
        self.assertTrue(ck["E4_cfo_direct_indirect"]["ok"])
        self.assertTrue(ck["can_issue"])

    def test_e4_excludes_non_operating_disposal_gain(self):
        # 资产处置收益(标 investing)进净利，但对应现金在投资活动 → 间接法剔除非经营损益，E4 仍平。
        # （覆盖性验证发现：此前 E4 对资产处置误报不平；剔除非经营损益 + 折旧取计提额 后恒等成立。）
        self._build()
        ledger.post_manual_entry([
            {"account": "1002 银行存款 Bank", "debit": "1200"},
            {"account": "1509 累计折旧 Accumulated Depreciation", "debit": "800"},
            {"account": "1500 固定资产 Fixed Assets", "credit": "1800"},
            {"account": "4200 营业外收入 Other Income", "credit": "200"}],
            date="2026-06-08", activity="investing")
        ind = reports.cash_flow_indirect()
        self.assertEqual(D(ind["non_operating_pl_excluded"]), D("200"))   # 处置收益被剔除
        self.assertEqual(D(ind["depreciation_addback"]), D("800"))        # 折旧取计提额，不被处置冲销抵消
        self.assertTrue(ind["e4_ok"])                                     # 直接法 == 间接法仍成立
        self.assertTrue(reports.checks()["can_issue"])


if __name__ == "__main__":
    unittest.main()
